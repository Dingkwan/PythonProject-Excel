from types import NoneType
from openpyxl import load_workbook

book=load_workbook("下单案件统计20220114.xlsx")

sheet=book["Sheet5"]

max_rows=sheet.max_row				#表格的最大行数


#获取下单编号
def getOrderNo():
	orderList=[]
	for rows in range(1,max_rows+1):
		val=sheet["C%d" %rows].value
		if (type(val)==NoneType):
			continue
		orderList.append(val)
	orderList.remove("下单编号")
	return orderList

#获取合同金额
def getPrice():
	priceList=[]
	for rows in range(1,max_rows+1):
		val=sheet["A%d" %rows].value
		if (type(val)==NoneType):		#去除空单元格
			continue
		if ("合同金额" in val):
			priceList.append(val)		#筛选合同金额
	return priceList


#获取客户名称
def getCustomers():
	customersList=[]
	for rows in range(1,max_rows+1):
		val=sheet["A%d" %rows].value
		if (type(val)==NoneType):			#去除空单元格
			continue
		if ("客户名称" in val):
			customersList.append(val)		#筛选客户名称
	return customersList


#获取下单名称
def getCases():
	casesList=[]
	for rows in range(1,max_rows+1):
		val=sheet["D%d" %rows].value
		if (type(val)==NoneType):
			continue
		casesList.append(val)
	casesList.remove("下单名称")
	return casesList


#获取下单状态 
def getOrderSatus():
	orderStausList=[]
	for rows in range(1,max_rows+1):
		val=sheet["I%d" %rows].value
		if (type(val)==NoneType):
			continue
		orderStausList.append(val)
	orderStausList.remove("审批状态")
	return orderStausList

#获取SE/SM/Sol
def getSE():
	SEList=[]
	for rows in range(1,max_rows+1):
		val=sheet["K%d" %rows].value
		if (type(val)==NoneType):
			continue
		SEList.append(val)
	SEList.remove("SE/SM/Sol推进")
	return SEList


#获取AM
def getAM():
	AMList=[]
	for rows in range(1,max_rows+1):
		val=sheet["J%d" %rows].value
		if (type(val)==NoneType):
			continue
		AMList.append(val)
	AMList.remove("AM")
	return AMList

#获取询价编号
def getPriceNo():
	priceNoList=[]
	for rows in range(1,max_rows+1):
		val=sheet["E%d" %rows].value
		if (type(val)==NoneType):
			continue
		priceNoList.append(val)
	priceNoList.remove("询价编号")
	return priceNoList


#获取开始时间
def getStartTime():
	startTimeList=[]
	for rows in range(1,max_rows+1):
		val=sheet["A%d" %rows].value
		if (type(val)==NoneType):
			continue
		if ("创建时间" in val):
			startTimeList.append(val)
	return startTimeList

#获取结束时间
def getEndTime():
	startEndList=[]
	for rows in range(1,max_rows+1):
		val=sheet["A%d" %rows].value
		if (type(val)==NoneType):
			continue
		if ("结束时间" in val):
			startEndList.append(val)
	return startEndList